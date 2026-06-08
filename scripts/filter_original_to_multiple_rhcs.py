import io
from getpass import getpass
from pathlib import Path

import msoffcrypto
from openpyxl import load_workbook


project_dir = Path(__file__).resolve().parent.parent
original_path = project_dir / "input files" / "Sequential RV-MPS in PAH (DS Copy).xlsx"
updated_path = project_dir / "output files" / "Sequential RV-MPS in PAH (DS Copy - Updated).xlsx"
output_path = project_dir / "output files" / "Sequential RV-MPS in PAH (DS Copy - Multiple RHCs Only).xlsx"

password = getpass("Enter Excel password: ")


def decrypt_file(path, password):
    buf = io.BytesIO()
    with open(path, "rb") as f:
        office = msoffcrypto.OfficeFile(f)
        office.load_key(password=password)
        office.decrypt(buf)
    buf.seek(0)
    return buf


def encrypt_workbook(unencrypted_buf, password):
    unencrypted_buf.seek(0)
    encrypted_buf = io.BytesIO()
    file_to_encrypt = msoffcrypto.OfficeFile(unencrypted_buf)
    file_to_encrypt.load_key(password=password)
    file_to_encrypt.encrypt(password, encrypted_buf)
    encrypted_buf.seek(0)
    return encrypted_buf


def clean_patient_id(value):
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    if text.endswith(".0"):
        text = text[:-2]
    return text


print("Loading original workbook...")
original_buf = decrypt_file(original_path, password)
original_wb = load_workbook(original_buf)
original_ws = original_wb.worksheets[0]

print("Loading updated workbook...")
updated_buf = decrypt_file(updated_path, password)
updated_wb = load_workbook(updated_buf, read_only=True, data_only=True)
updated_ws = updated_wb.worksheets[0]

# Patient ID is expected to be in the first column of both sheets.
patients_to_keep = {
    clean_patient_id(row[0])
    for row in updated_ws.iter_rows(min_row=2, max_col=1, values_only=True)
}
patients_to_keep.discard(None)
updated_wb.close()

print(f"Found {len(patients_to_keep)} patients to keep from the updated workbook")

deleted_rows = 0
kept_rows = 0

# Delete from bottom to top so row numbers stay valid while deleting.
for row_idx in range(original_ws.max_row, 1, -1):
    patient_id = clean_patient_id(original_ws.cell(row=row_idx, column=1).value)
    if patient_id in patients_to_keep:
        kept_rows += 1
    else:
        original_ws.delete_rows(row_idx)
        deleted_rows += 1

print(f"Kept {kept_rows} patient rows")
print(f"Deleted {deleted_rows} patient rows")

unencrypted_out = io.BytesIO()
original_wb.save(unencrypted_out)
original_wb.close()

encrypted_out = encrypt_workbook(unencrypted_out, password)

with open(output_path, "wb") as f:
    f.write(encrypted_out.getvalue())

del password
print(f"Done. Saved to: {output_path}")
